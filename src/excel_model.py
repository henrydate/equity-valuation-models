"""
Excel model generator
======================

Renders a :class:`~src.ledger.Ledger` + :class:`~src.sotp.SumOfParts` into a
formatted ``.xlsx`` workbook -- the co-primary deliverable alongside the
markdown note. Sheets:

- **Cover** -- company, recommendation, headline target & WACC, disclaimer.
- **Valuation** -- the sum-of-parts bridge to equity / per share.
- **WACC** -- the discount-rate build.
- **Assumptions** -- the ledger, colour-coded by verification (the provenance trail).
- **Sensitivity** -- per-share vs WACC x commodity-price shift, plus a
  bear/base/bull scenario chart.

Pure openpyxl; no template files. All figures inherit the ledger's
illustrative/unverified status -- the workbook is the model, not advice.
"""

from __future__ import annotations

from datetime import date
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .ledger import InputKind, Ledger, VerificationStatus

NAVY = "1F3864"
_TITLE = Font(bold=True, size=16, color=NAVY)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_FILL_HEADER = PatternFill("solid", fgColor=NAVY)
_FILL_VERIFIED = PatternFill("solid", fgColor="C6EFCE")    # green
_FILL_UNVERIFIED = PatternFill("solid", fgColor="FFEB9C")  # amber
_FILL_TOTAL = PatternFill("solid", fgColor="D9E1F2")       # light blue
_CENTER = Alignment(horizontal="center")
_WRAP = Alignment(wrap_text=True, vertical="top")
_MONEY = "#,##0"
_PX = "#,##0.00"
_PCT = "0.0%"


def _header(ws, row: int, headers: List[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _FILL_HEADER
        c.alignment = _CENTER


def _widths(ws, widths) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _banner(ws, row: int, text: str, span: int = 6) -> None:
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color="9C0006")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _cover(wb: Workbook, ledger: Ledger, sotp, recommendation: Optional[str]) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = f"{ledger.company} ({ledger.ticker})"
    ws["A1"].font = _TITLE
    ws["A2"] = f"Equity valuation model  ·  prepared {date.today().isoformat()}"
    _banner(ws, 4, "ILLUSTRATIVE -- NOT INVESTMENT ADVICE. Figures are placeholders; "
                   "assumptions unverified. See the Assumptions sheet.")
    rows = [
        ("Rating", recommendation or "-"),
        ("Target / share (model ccy)", round(sotp.value_per_share(), 2)),
        ("Reporting currency", ledger.reporting_currency),
        ("WACC", ledger.results.get("wacc", {}).get("value")),
        ("Enterprise value (m)", round(sotp.enterprise_value())),
        ("Equity value (m)", round(sotp.equity_value())),
        ("Entries (verified / total)",
         f"{ledger.audit_summary()['verified']} / {ledger.audit_summary()['entries_total']}"),
    ]
    for i, (k, v) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=k).font = _BOLD
        cell = ws.cell(row=i, column=2, value=v)
        if k == "WACC" and isinstance(v, float):
            cell.number_format = _PCT
    _widths(ws, {"A": 30, "B": 26})


def _valuation(wb: Workbook, ledger: Ledger, sotp) -> None:
    ws = wb.create_sheet("Valuation")
    ws["A1"] = "Sum-of-the-parts valuation"
    ws["A1"].font = _TITLE
    _header(ws, 3, [f"Component ({ledger.reporting_currency}m)", "Value", "% of EV"])
    table = sotp.summary_table()
    r = 4
    totals = {"Enterprise value", "Equity value", "Value per share"}
    for comp, row in table.iterrows():
        ws.cell(row=r, column=1, value=comp)
        vcell = ws.cell(row=r, column=2, value=round(float(row["Value"]), 2))
        vcell.number_format = _PX if comp == "Value per share" else _MONEY
        pct = row["% of EV"]
        if pct == pct:  # not NaN
            pcell = ws.cell(row=r, column=3, value=float(pct))
            pcell.number_format = _PCT
        if comp in totals:
            for col in (1, 2, 3):
                ws.cell(row=r, column=col).font = _BOLD
                ws.cell(row=r, column=col).fill = _FILL_TOTAL
        r += 1
    _widths(ws, {"A": 26, "B": 16, "C": 10})


def _wacc(wb: Workbook, ledger: Ledger) -> None:
    ws = wb.create_sheet("WACC")
    ws["A1"] = "WACC build"
    ws["A1"].font = _TITLE
    _header(ws, 3, ["Component", "Value"])
    g = ledger.value_of
    res = ledger.results
    items = [
        ("Risk-free rate", g("wacc.risk_free"), _PCT),
        ("Equity risk premium", g("wacc.erp"), _PCT),
        ("Equity beta", g("wacc.equity_beta"), "0.00"),
        ("Cost of equity (CAPM)", res.get("cost_of_equity", {}).get("value"), _PCT),
        ("Pre-tax cost of debt", g("wacc.cost_of_debt"), _PCT),
        ("Tax rate", g("wacc.tax_rate"), _PCT),
        ("Market value of equity", g("wacc.mv_equity"), _MONEY),
        ("Market value of net debt", g("wacc.mv_debt"), _MONEY),
        ("WACC", res.get("wacc", {}).get("value"), _PCT),
    ]
    r = 4
    for label, val, fmt in items:
        ws.cell(row=r, column=1, value=label)
        if val is not None:
            c = ws.cell(row=r, column=2, value=float(val))
            c.number_format = fmt
        if label in ("Cost of equity (CAPM)", "WACC"):
            ws.cell(row=r, column=1).font = _BOLD
            ws.cell(row=r, column=2).font = _BOLD
            ws.cell(row=r, column=1).fill = _FILL_TOTAL
            ws.cell(row=r, column=2).fill = _FILL_TOTAL
        r += 1
    _widths(ws, {"A": 26, "B": 16})


def _assumptions(wb: Workbook, ledger: Ledger) -> None:
    ws = wb.create_sheet("Assumptions")
    ws["A1"] = "Assumptions & provenance (the ledger)"
    ws["A1"].font = _TITLE
    headers = ["Key", "Input", "Value", "Unit", "Kind", "Source", "Verified", "Method", "Rationale", "Citation"]
    _header(ws, 3, headers)
    r = 4
    for key, e in ledger.entries.items():
        value = "; ".join(f"{k}:{v}" for k, v in e.value.items()) if isinstance(e.value, dict) else e.value
        verified = e.verification == VerificationStatus.VERIFIED
        rowvals = [key, e.label, value, e.unit, e.kind.value,
                   e.source_type.name.replace("_", " ").title(),
                   "yes" if verified else "no", e.method or "",
                   e.rationale or "", e.citation]
        for i, v in enumerate(rowvals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.fill = _FILL_VERIFIED if verified else _FILL_UNVERIFIED
            if headers[i - 1] in ("Rationale", "Citation"):
                c.alignment = _WRAP
        r += 1
    _widths(ws, {"A": 22, "B": 26, "C": 22, "D": 8, "E": 13, "F": 16,
                 "G": 9, "H": 12, "I": 44, "J": 30})


def _sensitivity(wb: Workbook, ledger: Ledger, sotp) -> None:
    ws = wb.create_sheet("Sensitivity")
    ws["A1"] = "Sensitivity -- per share"
    ws["A1"].font = _TITLE
    ws["A2"] = "rows: WACC   ·   cols: uniform commodity-price shift"

    base_wacc = sotp.discount_rate
    wacc_steps = [base_wacc - 0.01, base_wacc - 0.005, base_wacc, base_wacc + 0.005, base_wacc + 0.01]
    price_factors = [0.85, 0.925, 1.0, 1.075, 1.15]

    ws.cell(row=4, column=1, value="WACC \\ price").font = _BOLD
    for j, pf in enumerate(price_factors, start=2):
        c = ws.cell(row=4, column=j, value=pf - 1)  # show as +/- %
        c.number_format = _PCT
        c.font = _HEADER_FONT
        c.fill = _FILL_HEADER
    for i, w in enumerate(wacc_steps, start=5):
        wc = ws.cell(row=i, column=1, value=w)
        wc.number_format = _PCT
        wc.font = _HEADER_FONT
        wc.fill = _FILL_HEADER
        for j, pf in enumerate(price_factors, start=2):
            v = sotp.revalue(discount_rate=w, price_factor=pf)
            cell = ws.cell(row=i, column=j, value=round(v, 2))
            cell.number_format = _PX
            if abs(w - base_wacc) < 1e-9 and abs(pf - 1.0) < 1e-9:
                cell.fill = _FILL_TOTAL
                cell.font = _BOLD

    # bear / base / bull scenario block + chart
    base = sotp.value_per_share()
    scen_row = 12
    ws.cell(row=scen_row, column=1, value="Scenario").font = _BOLD
    ws.cell(row=scen_row, column=2, value="Per share").font = _BOLD
    scenarios = [
        ("Bear (-15% px, +1% WACC)", sotp.revalue(base_wacc + 0.01, 0.85)),
        ("Base", base),
        ("Bull (+15% px, -1% WACC)", sotp.revalue(base_wacc - 0.01, 1.15)),
    ]
    for k, (name, val) in enumerate(scenarios, start=scen_row + 1):
        ws.cell(row=k, column=1, value=name)
        ws.cell(row=k, column=2, value=round(val, 2)).number_format = _PX

    chart = BarChart()
    chart.title = "Valuation range (per share)"
    chart.type = "bar"
    chart.legend = None
    data = Reference(ws, min_col=2, min_row=scen_row + 1, max_row=scen_row + 3)
    cats = Reference(ws, min_col=1, min_row=scen_row + 1, max_row=scen_row + 3)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, "D12")
    _widths(ws, {"A": 24, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12})


def build_workbook(ledger: Ledger, sotp, *, recommendation: Optional[str] = None) -> Workbook:
    wb = Workbook()
    _cover(wb, ledger, sotp, recommendation)
    _valuation(wb, ledger, sotp)
    _wacc(wb, ledger)
    _assumptions(wb, ledger)
    _sensitivity(wb, ledger, sotp)
    return wb


def write_workbook(path, ledger: Ledger, sotp, **kwargs) -> str:
    wb = build_workbook(ledger, sotp, **kwargs)
    wb.save(path)
    return str(path)
