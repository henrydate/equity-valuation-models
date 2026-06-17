from openpyxl import load_workbook

from src.ledger import Ledger, LedgerEntry, InputKind, VerificationStatus
from src.financial_statements import SourceOfTruth
from src.sotp import AssetValuation, Division, SumOfParts
from src.excel_model import build_workbook, write_workbook


def build_ledger():
    led = Ledger("BHP Group Limited", "BHP.AX")
    led.add(LedgerEntry(key="wacc.risk_free", label="Risk-free rate", value=0.042, unit="%",
        kind=InputKind.DISCRETIONARY, source_type=SourceOfTruth.EXTERNAL_DATA, citation="FRED",
        as_of="2026-06-15", verification=VerificationStatus.VERIFIED, rationale="USD basis"))
    led.add(LedgerEntry(key="group.share_price", label="Share price", value=41.2, unit="AUD",
        kind=InputKind.HARD_FACT, source_type=SourceOfTruth.EXTERNAL_DATA, citation="yfinance",
        as_of="2026-06-15", verification=VerificationStatus.UNVERIFIED, provenance_method="auto_pull"))
    led.set_result("wacc", 0.085, unit="%")
    led.set_result("cost_of_equity", 0.0876, unit="%")
    return led


def make_sotp():
    a = AssetValuation(name="Mine", commodity="iron ore",
                       production={2027: 100, 2028: 100}, price={2027: 90, 2028: 88},
                       unit_cash_cost={2027: 20, 2028: 20}, tax_rate=0.30)
    return SumOfParts(company="BHP", base_year=2026, discount_rate=0.085,
                      divisions=[Division("Iron Ore", [a])], net_debt=1000, shares_outstanding=100)


def _find_value_right_of(ws, label):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                return ws.cell(row=cell.row, column=cell.column + 1).value
    return None


def test_workbook_has_all_sheets(tmp_path):
    path = tmp_path / "model.xlsx"
    write_workbook(path, build_ledger(), make_sotp(), recommendation="HOLD")
    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Cover", "Valuation", "WACC", "Assumptions", "Sensitivity"}


def test_valuation_per_share_matches(tmp_path):
    sotp = make_sotp()
    path = tmp_path / "model.xlsx"
    write_workbook(path, build_ledger(), sotp)
    wb = load_workbook(path)
    cell = _find_value_right_of(wb["Valuation"], "Value per share")
    assert abs(cell - round(sotp.value_per_share(), 2)) < 0.01


def test_assumptions_has_a_row_per_entry(tmp_path):
    led = build_ledger()
    path = tmp_path / "model.xlsx"
    write_workbook(path, led, make_sotp())
    wb = load_workbook(path)
    ws = wb["Assumptions"]
    # header on row 3, then one row per entry
    data_rows = [r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=1).value]
    assert len(data_rows) == len(led.entries)


def test_sensitivity_centre_equals_base(tmp_path):
    sotp = make_sotp()
    path = tmp_path / "model.xlsx"
    write_workbook(path, build_ledger(), sotp)
    wb = load_workbook(path)
    ws = wb["Sensitivity"]
    # centre cell of the 5x5 grid (row 7, col 4) == base per share
    centre = ws.cell(row=7, column=4).value
    assert abs(centre - round(sotp.value_per_share(), 2)) < 0.01
